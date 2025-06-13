import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
from datetime import datetime, timedelta

# --- SUA LISTA DE APARELHOS E PATRIMÔNIOS ---
APARELHOS_DA_ACADEMIA = {
    "Glúteo Máquina": "1352348",
    "Cadeira extensora 1": "1352356",
    "Cadeira extensora 2": "1352355",
    "Cadeira abdutora 1": "1352363",
    "Cadeira abdutora 2": "1352362",
    "Cadeira adutora 1": "1352364",
    "Cadeira adutora 2": "1352361",
    "Panturrilha em pé": "1352344",
    "Panturrilha maquina sentado 1": "1339634",
    "Panturrilha maquina sentado 2": "1374039",
    "Barra guiada": "1352345",
    "Barra livre": "1339635",
    "Leg press maquina": "1352352",
    "Cadeira flexora 1": "1352353",
    "Cadeira flexora 2": "1352354",
    "Leg 45 1": "1352365",
    "Leg 45 2": "1352366",
    "Tríceps maquina": "1352340",
    "Peck Deck": "1352347",
    "Supino reto máquina": "1352343",
    "Desenvolvimento maquina": "1352351",
    "Puxador frente aberta 1": "1352359",
    "Puxador frente aberta 2": "1352360",
    "Graviton": "1352346",
    "Remada maquina 1": "1170772",
    "Remada maquina 2": "1352341",
    "Crossover 1": "1352358",
    "Crossover 2": "1352357",
    "Lombar maquina": "1339640",
    "Rosca Scott Maquina": "1339636",
    "Abdominal máquina 1": "1574225",
    "Abdominal máquina 2": "1352350",
    "Banco regulável 1": "1339638",
    "Banco regulável 2": "1339637",
    "Banco supino 1": "1339642",
    "Banco supino 2": "1339641",
    "Elevação pélvica": "0"
}

def get_or_create_sheet(workbook, sheet_name, headers):
    """
    Verifica se uma aba existe no workbook, se não, a cria com os cabeçalhos.
    Retorna a aba (sheet) carregada ou recém-criada.
    """
    clean_sheet_name = sheet_name
    for char in ['/', '\\', '?', '*', '[', ']', ':']:
        clean_sheet_name = clean_sheet_name.replace(char, '_')
    clean_sheet_name = clean_sheet_name[:31]

    if clean_sheet_name not in workbook.sheetnames:
        print(f"    - Criando nova aba para '{clean_sheet_name}'...")
        current_sheet = workbook.create_sheet(title=clean_sheet_name)
        current_sheet.append(headers)
        for i, header in enumerate(headers):
            current_sheet.column_dimensions[get_column_letter(i + 1)].width = len(header) + 5
    else:
        current_sheet = workbook[clean_sheet_name]
    return current_sheet

def find_last_preventiva_date(sheet):
    """
    Percorre a aba de um aparelho para encontrar a data da última manutenção preventiva.
    Retorna a data como string DD/MM/AAAA ou None se nenhuma preventiva for encontrada.
    """
    # A coluna 'Tipo de Manutenção Realizada' é a coluna 1 (índice 0) dos cabeçalhos do aparelho
    # A coluna 'Data da Manutenção' é a coluna 2 (índice 1) dos cabeçalhos do aparelho

    # Percorre as linhas de baixo para cima para encontrar a última preventiva mais rapidamente
    for row_idx in range(sheet.max_row, 1, -1): # Começa da última linha, vai até a linha 2 (ignora cabeçalho)
        row_values = [cell.value for cell in sheet[row_idx]]

        # Garante que a linha tenha dados suficientes para não dar erro de índice
        if len(row_values) > 1 and row_values[0] == "Preventivo":
            data_str = str(row_values[2]) # Assumindo que Data da Manutenção é o 3º item (índice 2)
            try:
                # Tenta converter para datetime para validar o formato
                datetime.strptime(data_str, "%d/%m/%Y")
                return data_str
            except ValueError:
                # Se não for uma data válida no formato, continua procurando
                pass
    return None # Nenhuma preventiva encontrada

def registrar_manutencao():
    """
    Função principal para registrar manutenções, com escolha entre preventiva e corretiva.
    """
    nome_arquivo = 'preventivas_academia.xlsx'

    if not os.path.exists(nome_arquivo):
        print(f"Arquivo '{nome_arquivo}' não encontrado. Criando um novo...")
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        main_sheet_headers = [
            "Nome do Aparelho",
            "Número de Patrimônio",
            "Tipo de Manutenção Realizada",
            "Responsável pela Manutenção",
            "Data da Manutenção",
            "Próxima Manutenção Prevista", # Agora será "Data da Última Corretiva" para corretivas
            "Observações",
            "Status"
        ]
        main_sheet = get_or_create_sheet(workbook, "Resumo Geral de Manutenções", main_sheet_headers)

        workbook.save(nome_arquivo)
        print(f"Arquivo '{nome_arquivo}' criado com sucesso e Planilha Mestra adicionada.")
    else:
        workbook = openpyxl.load_workbook(nome_arquivo)
        print(f"Arquivo '{nome_arquivo}' carregado.")

    print("\n--- Registro de Manutenção ---")
    while True:
        tipo_manutencao_escolha = input("Qual o tipo de manutenção a ser registrada? (P)reventiva ou (C)orretiva? ").strip().lower()
        if tipo_manutencao_escolha == 'p':
            executar_preventiva(workbook)
            break
        elif tipo_manutencao_escolha == 'c':
            executar_corretiva(workbook)
            break
        else:
            print("Opção inválida. Por favor, digite 'P' para Preventiva ou 'C' para Corretiva.")

    workbook.save(nome_arquivo)
    print(f"\nOperação concluída. O arquivo '{nome_arquivo}' foi atualizado.")

def executar_preventiva(workbook):
    """
    Executa o fluxo de registro de manutenção preventiva para todos os aparelhos.
    """
    print("\n--- Registrando Manutenção Preventiva Geral ---")
    responsavel = input("Responsável pela Manutenção Preventiva Geral: ")

    observacoes_especificas = {}
    while True:
        has_observation = input("Há alguma observação para algum aparelho específico? (s/n): ").strip().lower()
        if has_observation == 's':
            print("\n--- Adicionar Observação Específica para Preventiva ---")
            aparelhos_list = list(APARELHOS_DA_ACADEMIA.keys())
            for i, aparelho_nome in enumerate(aparelhos_list):
                print(f"{i+1}. {aparelho_nome} (Patrimônio: {APARELHOS_DA_ACADEMIA[aparelho_nome]})")

            selected_aparelho_name = ""
            while True:
                try:
                    choice = input("Digite o NÚMERO do aparelho para adicionar a observação (ou 'f' para finalizar a adição de observações): ").strip().lower()
                    if choice == 'f':
                        break

                    index = int(choice) - 1
                    if 0 <= index < len(aparelhos_list):
                        selected_aparelho_name = aparelhos_list[index]
                        observacao_text = input(f"Digite a observação para '{selected_aparelho_name}': ")
                        observacoes_especificas[selected_aparelho_name] = observacao_text
                        print(f"Observação adicionada para {selected_aparelho_name}.")
                        break
                    else:
                        print("Número inválido. Por favor, digite um número da lista.")
                except ValueError:
                    print("Entrada inválida. Por favor, digite um número ou 'f'.")

            if choice == 'f':
                break

        elif has_observation == 'n':
            print("Nenhuma observação específica será adicionada para a preventiva.")
            break
        else:
            print("Resposta inválida. Por favor, digite 's' para sim ou 'n' para não.")

    tipo_manutencao_registro = "Preventivo"
    data_manutencao = datetime.now().strftime("%d/%m/%Y")
    proxima_manutencao_data_obj = datetime.now() + timedelta(days=15) # Quinzenal
    proxima_manutencao = proxima_manutencao_data_obj.strftime("%d/%m/%Y")
    status = "Concluída"

    print("\nRegistrando manutenções preventivas para todos os aparelhos...")
    main_sheet = workbook["Resumo Geral de Manutenções"]
    device_sheet_headers = [
        "Tipo de Manutenção Realizada",
        "Responsável pela Manutenção",
        "Data da Manutenção",
        "Próxima Manutenção Prevista",
        "Observações",
        "Status"
    ]

    for nome_aparelho, numero_patrimonio in APARELHOS_DA_ACADEMIA.items():
        print(f"  - Registrando para: {nome_aparelho} (Patrimônio: {numero_patrimonio})")

        observacao_para_aparelho = observacoes_especificas.get(nome_aparelho, "")

        current_sheet = get_or_create_sheet(workbook, f"{nome_aparelho} ({numero_patrimonio})", device_sheet_headers)

        dados_nova_linha_aparelho = [
            tipo_manutencao_registro,
            responsavel,
            data_manutencao,
            proxima_manutencao,
            observacao_para_aparelho,
            status
        ]
        current_sheet.append(dados_nova_linha_aparelho)

        dados_nova_linha_mestra = [
            nome_aparelho,
            numero_patrimonio,
            tipo_manutencao_registro,
            responsavel,
            data_manutencao,
            proxima_manutencao,
            observacao_para_aparelho,
            status
        ]
        main_sheet.append(dados_nova_linha_mestra)

    print("\nManutenção preventiva geral registrada com sucesso!")

def executar_corretiva(workbook):
    """
    Executa o fluxo de registro de manutenção corretiva.
    """
    print("\n--- Registrando Manutenção Corretiva ---")

    device_sheet_headers = [
        "Tipo de Manutenção Realizada",
        "Responsável pela Manutenção",
        "Data da Manutenção",
        "Próxima Manutenção Prevista", # Onde ficará a data da última preventiva
        "Observações",
        "Status"
    ]

    main_sheet = workbook["Resumo Geral de Manutenções"]

    while True:
        aparelhos_list = list(APARELHOS_DA_ACADEMIA.keys())
        print("\n--- Selecione o aparelho para a Manutenção Corretiva ---")
        for i, aparelho_nome in enumerate(aparelhos_list):
            print(f"{i+1}. {aparelho_nome} (Patrimônio: {APARELHOS_DA_ACADEMIA[aparelho_nome]})")

        selected_aparelho_name = ""
        selected_patrimonio = ""
        while True:
            try:
                choice = input("Digite o NÚMERO do aparelho na lista: ")
                index = int(choice) - 1
                if 0 <= index < len(aparelhos_list):
                    selected_aparelho_name = aparelhos_list[index]
                    selected_patrimonio = APARELHOS_DA_ACADEMIA[selected_aparelho_name]
                    print(f"Você selecionou: {selected_aparelho_name} (Patrimônio: {selected_patrimonio})")
                    break
                else:
                    print("Número inválido. Por favor, digite um número da lista.")
            except ValueError:
                print("Entrada inválida. Por favor, digite um número.")

        responsavel = input(f"Responsável pela manutenção corretiva em '{selected_aparelho_name}': ")
        observacoes_corretiva = ""
        while not observacoes_corretiva:
            observacoes_corretiva = input(f"Observação DETALHADA sobre a manutenção corretiva em '{selected_aparelho_name}': ")
            if not observacoes_corretiva:
                print("A observação é obrigatória para manutenções corretivas. Por favor, detalhe o que foi corrigido.")

        tipo_manutencao_registro = "Corretiva"
        data_manutencao = datetime.now().strftime("%d/%m/%Y")
        status = "Concluída"

        # Carrega ou cria a aba específica do aparelho
        current_sheet = get_or_create_sheet(workbook, f"{selected_aparelho_name} ({selected_patrimonio})", device_sheet_headers)

        # --- LÓGICA PARA BUSCAR A ÚLTIMA PREVENTIVA E CALCULAR A PRÓXIMA DATA ---
        last_preventiva_date_str = find_last_preventiva_date(current_sheet)

        if last_preventiva_date_str:
            try:
                last_preventiva_date_obj = datetime.strptime(last_preventiva_date_str, "%d/%m/%Y")
                # A próxima preventiva é 15 dias após a última preventiva registrada
                proxima_manutencao = (last_preventiva_date_obj + timedelta(days=15)).strftime("%d/%m/%Y")
                print(f"    - Próxima preventiva agendada para: {proxima_manutencao}")
            except ValueError:
                proxima_manutencao = "Erro no cálculo da preventiva" # Caso a data lida esteja em formato inválido
                print(f"    - Erro ao calcular a próxima preventiva. Campo 'Próxima Manutenção Prevista' preenchido como '{proxima_manutencao}'.")
        else:
            proxima_manutencao = "Nenhuma preventiva anterior" # Se não houver preventivas registradas
            print(f"    - Nenhuma manutenção preventiva anterior encontrada para este aparelho. Campo 'Próxima Manutenção Prevista' preenchido como '{proxima_manutencao}'.")

        # Dados para preencher a linha
        dados_nova_linha_aparelho = [
            tipo_manutencao_registro,
            responsavel,
            data_manutencao,
            proxima_manutencao, # Data da próxima preventiva calculada
            observacoes_corretiva,
            status
        ]
        current_sheet.append(dados_nova_linha_aparelho)

        dados_nova_linha_mestra = [
            selected_aparelho_name,
            selected_patrimonio,
            tipo_manutencao_registro,
            responsavel,
            data_manutencao,
            proxima_manutencao, # Data da próxima preventiva calculada
            observacoes_corretiva,
            status
        ]
        main_sheet.append(dados_nova_linha_mestra)

        print(f"Manutenção corretiva para '{selected_aparelho_name}' registrada com sucesso.")

        registrar_outra = input("Deseja registrar outra manutenção corretiva? (s/n): ").strip().lower()
        if registrar_outra != 's':
            break

    print("\nRegistro de manutenções corretivas finalizado.")

# Chamada principal para executar o script
if __name__ == "__main__":
    registrar_manutencao()
